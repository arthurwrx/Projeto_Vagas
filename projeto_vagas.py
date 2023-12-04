""" 1 Perguntar o nome da empresa que deseja que busque
    2 Iniciar o navegador e acessar o site da Gupy
    3 Iterar sobre as vagas e a cada vaga salvar as informações nas colunas do Excel
"""
##Dependências
import PySimpleGUI as sg 
from botcity.web import WebBot, Browser
from botcity.web import By
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import datetime
import time

### Variáveis de Vagas de Sites:
#Gupy
nome_das_vagas = []
localidades_das_vagas = []
tipos_vagas = []

#Linkedin
nome_das_vagas_linkedin = []
localidades_das_vagas_linkedin = []
num_vagas = 0
nome_das_vagas_linkedin_selector = []
data_publicacoes_linkedin = []

#Glassdoor

nome_das_vagas_glassdoor = []
localidades_das_vagas_glassdoor = []
salario_das_vagas_glassdoor = []

#Outras Variáveis

bot = WebBot()
hoje = str(datetime.datetime.now().strftime('%d.%m.%Y %Hh%Mm'))

##Controle da tela
def tela_inicial():


    if __name__ == "__main__":


        sg.change_look_and_feel('Gray Gray Gray')

        tamanho_botao = (15,2)


    
        layout = [
            [sg.Column([[sg.Image(r'C:\Projetos Python\Projeto_Vagas\Logo\logo-assinatura.png')]], justification='center')],
            [sg.Column([[sg.Text('Bem vindo a Automação de Buscas de Vagas Abertas!',font=('Helvetica', 12 ,'bold'))]], justification='center')],
            [sg.Column([[sg.Text('Antes de começar, digite o nome da empresa desejada:',font=('Helvetica', 10, 'bold'))]], justification='center')],
            [sg.Column([[sg.Text('Dicas',font=('Helvetica', 10, 'bold'))]], justification='center')],
            [sg.Column([[sg.Text('01 - Escreva o nome da empresa idêntica a forma que está no Linkedin, \nvocê terá uma pesquisa mais precisa',font=('Helvetica', 10, 'bold'))]], justification='center')],
            [sg.Column([[sg.Text('02 - Verifique se a empresa que você quer existe na Gupy antes de iniciar',font=('Helvetica', 10, 'bold'))]], justification='center')],
            [sg.Text('Empresa desejada: '), sg.InputText(key="nome_empresa")],
            [sg.Radio('Gupy', key='gupy', default=True,group_id='plataforma')],
            [sg.Radio('Linkedin', key='linkedin',group_id='plataforma')],
            [sg.Radio('Glassdoor', key='glassdoor',group_id='plataforma')],
            [sg.Radio('Todas', key='todas',group_id='plataforma')],
            [sg.Column([[sg.Text('Se você já escolheu a empresa, clique em COMEÇAR! para prosseguir')]], justification='center')],
            [sg.Text('')],
            [
            sg.Column([[sg.Button('Começar!', size=tamanho_botao, font=('Helvetica', 10, 'bold'))]], justification='center', element_justification='center'),

            ]]
        

        window = sg.Window('CNPJ',layout, size=(700, 550))
        

        while True: 
            event, values = window.read()
            if event == sg.WIN_CLOSED:
                break

            elif event == 'Executar':

                if values['gupy']:
                    limpa_excel()
                    window.close()
                    nome_empresa = values['nome_empresa']
                    config_navegacao(nome_empresa)
                    captura_vagas()
                    joga_no_excel(nome_empresa)
                    tela_retorna_menu()

                if values['linkedin']:
                    limpa_excel()
                    window.close()
                    nome_empresa = values['nome_empresa']
                    busca_linkedin(nome_empresa)
                    joga_no_excel(nome_empresa)
                    tela_retorna_menu()
                
                if values['glassdoor']:
                    limpa_excel()
                    window.close()
                    nome_empresa = values['nome_empresa']
                    login_glassdoor(nome_empresa)


                if values['todas']:

                    limpa_excel()
                    window.close()
                    nome_empresa = values['nome_empresa']
                    config_navegacao(nome_empresa)
                    captura_vagas()
                    busca_linkedin(nome_empresa)
                    login_glassdoor(nome_empresa)


                

    return nome_empresa

## Função destinada a iniciar o navegador e entrar no site da gupy
def config_navegacao(nome_empresa):

    # Configure whether or not to run on headless mode.
    bot.headless = False
    bot.browser = Browser.CHROME
    bot.driver_path = "chromedriver.exe"
    download_folder_path = r'C:\Projetos Python\Projeto_Vagas\downloads'
    
    # Opens the browser on the BotCity website.
    bot.start_browser()
    bot.browse("https://www.google.com/")
                                      
    ## Abre o google
    pesquisa = bot.find_element("//textarea[@id='APjFqb']",By.XPATH)
    pesquisa.send_keys(f'{nome_empresa} gupy')
    bot.wait(3000)

    ## Pesquisa a empresa
    try:
        pesquisa_google_button = bot.find_element("//*[@value='Pesquisa Google']", By.XPATH)
        bot.wait(2000)
        pesquisa_google_button.click()
    
    except:
        pesquisa_google_button = bot.find_element("//div[@aria-label='Ferramentas de inserção de texto']",By.XPATH)
        pesquisa_google_button.click()
        bot.wait(2000)
        pesquisa_google_button = bot.find_element("html/body/div[1]/div[3]/form/div[1]/div[1]/div[4]/center/input[1]", By.XPATH)
        pesquisa_google_button.click()

    ##Site Gupy                  
    pesquisa = bot.find_element("//h3[@class='LC20lb MBeuO DKV0Md']",By.XPATH)
    pesquisa.click()
    
    try:
        gupy = bot.find_element("//button[contains(text(), 'Ok, entendi')]",By.XPATH)
        gupy.click()
    
    except:
        pass

    try:
        gupy = bot.find_element("//*[@id='onetrust-accept-btn-handler']",By.XPATH)
        bot.wait_for_element_visibility(element=gupy, visible=True, waiting_time=10000)
        gupy.click() 

    except:
        pass

## Função destinada a capturar os detalhes da vaga da empresa desejada
def captura_vagas():
   
    i = 1
    while i < 11:
        
        nome_vaga_selector = f"//*[@id='job-listing']/ul/li[{str(i)}]/a/div/div[1]"
        local_vaga_selector = f"//*[@id='job-listing']/ul/li[{str(i)}]/a/div/div[{2}]"
        tipo_vaga_selector = f"//*[@id='job-listing']/ul/li[{str(i)}]/a/div/div[{3}]"
        

        try:
            nome_vaga = bot.find_element(nome_vaga_selector,By.XPATH)
            nome_da_vaga_string = nome_vaga.text
            nome_das_vagas.append(nome_da_vaga_string)
            
            local_vaga = bot.find_element(local_vaga_selector,By.XPATH)
            local_da_vaga_string = local_vaga.text
            localidades_das_vagas.append(local_da_vaga_string)

            tipo_vaga = bot.find_element(tipo_vaga_selector,By.XPATH)
            tipo_da_vaga = tipo_vaga.text
            tipos_vagas.append(tipo_da_vaga)
            i += 1

        except NoSuchElementException:
            break  # Sai do loop while se o elemento não for encontrado
    
    try: 
        proxima_pagina = bot.find_element("//button[@aria-label='Vá para próxima página']",By.XPATH)
        if proxima_pagina.is_enabled():
            proxima_pagina.click()
            captura_vagas()

    except:
        print("capturamos tudo")

## Função destinada a dispor as informações no excel
def joga_no_excel(nome_empresa):

    workbook = openpyxl.Workbook()

    if nome_das_vagas != []:
        
        sheet_gupy = workbook.create_sheet(title="Gupy")
        sheet_gupy.append({'A': 'Nome da Vaga', 'B': 'Localidade', 'C': 'Tipo de Vaga'})
        
        for i in range(len(tipos_vagas)):
                sheet_gupy.append({
                    'A':nome_das_vagas[i],
                    'B':localidades_das_vagas[i],
                    'C':tipos_vagas[i]
                })
            
    if nome_das_vagas_linkedin != []:

        sheet_linkedin = workbook.create_sheet(title="Linkedin")
        sheet_linkedin.append({'A': 'Nome da Vaga', 'B': 'Localidade','C':'Quando foi publicada?'})


        for i in range(len(nome_das_vagas_linkedin)):
            sheet_linkedin.append({
            'A':nome_das_vagas_linkedin[i],
            'B':localidades_das_vagas_linkedin[i],
            'C':data_publicacoes_linkedin[i]
        })
            
    if nome_das_vagas_glassdoor != []:

        sheet_glassdoor = workbook.create_sheet(title="Glassdoor")
        sheet_glassdoor.append({'A': 'Nome da Vaga', 'B': 'Localidade'})
    
        for i in range(len(nome_das_vagas_glassdoor)):
                sheet_glassdoor.append({
                    'A':nome_das_vagas_glassdoor[i],
                    'B':localidades_das_vagas_glassdoor[i],
                })

     # Verifique se a planilha "Sheet" existe antes de tentar removê-la
    if "Sheet" in workbook.sheetnames:
        workbook.remove_sheet(workbook["Sheet"])

    workbook.save(filename=f"{nome_empresa} {hoje}.xlsx")
    
    bot.close_page()

## Função a limpar as listas do excel
def limpa_excel():
 
    global nome_das_vagas, localidades_das_vagas, tipos_vagas, nome_das_vagas_linkedin,localidades_das_vagas_linkedin,tipos_vagas_linkedin,email,senha
    nome_das_vagas = []
    localidades_das_vagas = []
    tipos_vagas = []

    nome_das_vagas_linkedin = []
    localidades_das_vagas_linkedin = []
    tipos_vagas_linkedin = []

## pergunta se o usuário quer fazer mais uma pesquisa
def tela_retorna_menu():

    if __name__ == "__main__":
        
        sg.change_look_and_feel('Gray Gray Gray')

        tamanho_botao = (15,2)
        tamanho_caixa = (10,5)

    
        layout = [

            [sg.Column([[sg.Image(r'C:\Projetos Python\Projeto_Vagas\Logo\logo-assinatura.png')]], justification='center')],
            [sg.Column([[sg.Text('Deseja consultar mais alguma empresa?',font=('Helvetica', 12 ,'bold'))]], justification='center')],
            [sg.Text('')],
            [
            sg.Column([[sg.Button('Sim', size=tamanho_botao, font=('Helvetica', 10, 'bold'))]], justification='center', element_justification='center'),
             sg.Column([[sg.Button('Não', size=tamanho_botao,font=('Helvetica', 10, 'bold'))]], justification='center', element_justification='center')]]
        

        window = sg.Window('CNPJ',layout, size=(700, 225))
        

        while True: 
            event, values = window.read()
            if event == sg.WIN_CLOSED:
                break

            elif event == 'Sim':
                window.close()
                tela_inicial()
                

            elif event == 'Não':
                window.close()
                break
            
def busca_linkedin(nome_empresa):
        
        
        bot.browse('https://www.linkedin.com/jobs/search?keywords=Cia%20de%20talentos&location=Brasil&geoId=106057199&trk=public_jobs_jobs-search-bar_search-submit&position=1&pageNum=0')
        bot.wait(3000)
        pesquisa_linkedin = bot.find_element('//*[@id="job-search-bar-keywords"]',By.XPATH)

        pesquisa_linkedin.clear()

        for char in nome_empresa:
            pesquisa_linkedin.send_keys(char)
            time.sleep(0.3)
        try:
            pesquisa_linkedin.click()
            pesquisa_linkedin.click()
            pesquisa_linkedin = bot.find_element("//*[@id='keywords-1']",By.XPATH)
            bot.wait(4000)
            empresa_buscada = pesquisa_linkedin.text
            pesquisa_linkedin.click()

        except:
            pass

        filtro_linkedin = bot.find_element("//button[@aria-label='Filtro Empresa. Clicar neste botão exibe todas as opções de filtro de Empresa.']",By.XPATH)
        filtro_linkedin.click()

        lista_filtros_linkedin = bot.find_elements("//div[@aria-label='Opções de filtro de Empresa.']/div[@class='filter-values-container__filter-value']",By.XPATH)
        
        ## Nesta etapa está ocorrendo a parte de filtragem da empresa

        for i in range(0,len(lista_filtros_linkedin)):

            empresa_filtrada = lista_filtros_linkedin[i]
            lista_checkbox_linkedin = f"//*[@id='f_C-{str(i)}']"
            empresa_filtrada = empresa_filtrada.text
            empresa_tratada = empresa_filtrada.split("(", 1)[0].strip()


            try:
                if empresa_tratada == empresa_buscada:
                    bot.wait(4000)
                    print(f"Condição satisfeita em: {empresa_tratada} {empresa_tratada}")
                    seleciona_caixinha = bot.find_element(lista_checkbox_linkedin,By.XPATH)
                    bot.wait(4000)
                    seleciona_caixinha.click()
            except:
                    if empresa_tratada == pesquisa_linkedin.text:
                        bot.wait(4000)
                        print(f"Condição satisfeita em: {empresa_tratada} {empresa_tratada}")
                        seleciona_caixinha = bot.find_element(lista_checkbox_linkedin,By.XPATH)
                        bot.wait(4000)
                        seleciona_caixinha.click()

        bot.wait(3000)
        clica_concluir = bot.find_element("//*[@id='jserp-filters']/ul/li[2]/div/div/div/button",By.XPATH)
        clica_concluir.click()
        bot.wait(5000)

        ### Começa aqui a etapa de capturar as vagas

        ## O número de scrolladas deve ser igual a o número de vagas dividido por 25 (Que é o número máximo de páginas abertas)

        num_vagas = f'//span[@class="results-context-header__job-count"]'
        num_vagas_site = bot.find_element(num_vagas,By.XPATH)
        num_vagas = num_vagas_site.text
        num_vagas = int(num_vagas)
        
        num_scrolls = num_vagas / 25
        num_scrolls = int(num_scrolls)

        for i in range((num_scrolls) + 1):

            try:
                bot.execute_javascript("window.scrollBy(0, document.body.scrollHeight);")
                bot.wait(1700)
                botao = bot.find_element("//button[@aria-label='Ver mais vagas']",By.XPATH)
                bot.wait(1700)
                botao.click()
                desce_pag += 1
            except:
                continue

        ## Nesta etapa, ocorre a captura das vagas e jogar na lista

        nome_das_vagas_linkedin_selector = bot.find_elements("//h3[@class='base-search-card__title']",By.XPATH)
        localizacao_atuacao_linkedin_selector = bot.find_elements("//span[@class='job-search-card__location']",By.XPATH)
        data_publi_linkedin_selector = bot.find_elements("//time[contains(@class, 'job-search-card__listdate')]",By.XPATH) 


        for i in range(len(nome_das_vagas_linkedin_selector)):

            ##Captura título da vaga
            nome_da_vaga_linkedin = nome_das_vagas_linkedin_selector[i]
            nome_da_vaga_linkedin = nome_da_vaga_linkedin.text
            nome_das_vagas_linkedin.append(nome_da_vaga_linkedin)
          

            ##Captura localidade
            localizacao_atuacao_linkedin = localizacao_atuacao_linkedin_selector[i]
            localizacao_atuacao_linkedin = localizacao_atuacao_linkedin.text
            localidades_das_vagas_linkedin.append(localizacao_atuacao_linkedin)

            ##Captura data de publicação da vaga
            data_publi_linkedin = data_publi_linkedin_selector[i]
            data_publi_linkedin = data_publi_linkedin.text
            data_publicacoes_linkedin.append(data_publi_linkedin)



def glassdoor_empregos(nome_empresa,email,senha):

        bot.browse('https://www.glassdoor.com.br/profile/login_input.htm')
        bot.wait(3000)

        login_gdempregos = bot.find_element("//*[@id='inlineUserEmail']",By.XPATH)
        login_gdempregos.send_keys(email)
        bot.wait(1000)
        bot.enter()

        login_gdempregos = bot.find_element("//*[@id='inlineUserPassword']",By.XPATH)
        bot.wait(2000)
        login_gdempregos.send_keys(senha)
        bot.wait(2000)
        bot.enter()

        pesquisa_gdempregos = bot.find_element('//*[@id="ContentNav"]/li[2]/a',By.XPATH)
        bot.wait(2000)
        pesquisa_gdempregos.click()

        pesquisa_gdempregos = bot.find_element("//input[@aria-label='Buscar empresa']",By.XPATH)
        pesquisa_gdempregos.send_keys(nome_empresa)

        pesquisa_gdempregos = bot.find_element('//*[@id="Explore"]/div[2]/div/div/div[2]/div/div/div/ul/li[1]',By.XPATH)
        bot.type_down()
        bot.enter()

        pesquisa_gdempregos = bot.find_element('//a[@data-test="ei-nav-jobs-link"]',By.XPATH)
        pesquisa_gdempregos.click()

def glassdoor_captura_vagas():

    nome_das_vagas_glassdoor_selector = bot.find_elements("//div[contains(@class, 'job-title') and contains(@id, 'job-title-')]",By.XPATH)
    localidades_das_vagas_glassdoor_selector = bot.find_elements("//div[contains(@class, 'location') and contains(@id, 'job-location-')]",By.XPATH)
    salario_das_vagas_glassdoor_selector = bot.find_elements("//div[@class='salary-estimate' and @data-test='detailSalary']",By.XPATH)

    for i in range(len(nome_das_vagas_glassdoor_selector)):

        ##Captura título da vaga
        nome_da_vaga_glassdoor = nome_das_vagas_glassdoor_selector[i]
        nome_da_vaga_glassdoor = nome_da_vaga_glassdoor.text
        nome_das_vagas_glassdoor.append(nome_da_vaga_glassdoor)
        

        ##Captura localidade da vaga
        localidade_da_vaga_glassdoor = localidades_das_vagas_glassdoor_selector[i]
        localidade_da_vaga_glassdoor = localidade_da_vaga_glassdoor.text
        localidades_das_vagas_glassdoor.append(localidade_da_vaga_glassdoor)

        # try:
        #     salario_glassdoor = salario_das_vagas_glassdoor_selector[i]
        #     salario_glassdoor = salario_glassdoor.text
        #     salario_das_vagas_glassdoor.append(salario_glassdoor)

        # except:
        #         salario_das_vagas_glassdoor.append("")
        
    try: 
        proxima_pagina_glassdoor = bot.find_element("//button[@aria-label='Next']",By.XPATH)
        if proxima_pagina_glassdoor.is_enabled():
            proxima_pagina_glassdoor.click()
            bot.key_esc()
            bot.key_esc()
            bot.key_esc()
            glassdoor_captura_vagas()
    except:
        pass

def login_glassdoor(nome_empresa):

    if __name__ == "__main__":
        
        sg.change_look_and_feel('Gray Gray Gray')

        tamanho_botao = (15,2)
    
        layout = [

            [sg.Column([[sg.Image(r'C:\Projetos Python\Projeto_Vagas\Logo\logo-assinatura.png')]], justification='center')],
            [sg.Column([[sg.Text('Digite seu email e senha do GlassDoor',font=('Helvetica', 12 ,'bold'))]], justification='center')],
            [sg.Text('')],
            [sg.Text('Email de Login: '), sg.InputText(key="login")],
            [sg.Text('Senha:             '), sg.InputText(key="senha")],
            [sg.Text('')],
            [sg.Text('')],
            [sg.Column([[sg.Button('Login', size=tamanho_botao, font=('Helvetica', 10, 'bold'))]], justification='center', element_justification='center')],
        ]

        window = sg.Window('Arthur',layout, size=(550, 350))
        

        while True: 
            event, values = window.read()
            
            if event == sg.WIN_CLOSED:
                break

            elif event == 'Login':
                window.close()
                email = values['login']
                senha = values['senha']
                glassdoor_empregos(nome_empresa,email,senha)
                glassdoor_captura_vagas()
                joga_no_excel(nome_empresa)
                tela_retorna_menu()

            
            else:
                window.close()
        
    return email,senha

tela_inicial()














    



